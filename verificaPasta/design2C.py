import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import PyPDF2
from datetime import datetime
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import threading
import json

class CNDManager:
    def __init__(self):
        # Configurações do CustomTkinter
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        
        # Janela principal
        self.root = ctk.CTk()
        self.root.title("CND Manager - Controle de Certidões")
        self.root.geometry("1000x700")
        self.root.minsize(900, 600)
        
        # Variáveis
        self.folder_path = tk.StringVar()
        self.processing = False
        self.results_data = []
        
        # Configurações padrão
        self.config = {
            "expected_files": ["CND MUNICIPAL", "CND RFB", "CND FGTS", "CND PROC", "CND ESTADUAL"],
            "target_line": "CERTIDÃO POSITIVA DE DÉBITOS - CPD",
            "last_folder": ""
        }
        
        self.load_config()
        self.create_widgets()
        self.center_window()
        
    def load_config(self):
        """Carrega configurações salvas"""
        try:
            if os.path.exists("cnd_config.json"):
                with open("cnd_config.json", "r", encoding="utf-8") as f:
                    loaded_config = json.load(f)
                    self.config.update(loaded_config)
        except Exception as e:
            print(f"Erro ao carregar configurações: {e}")
    
    def save_config(self):
        """Salva configurações"""
        try:
            with open("cnd_config.json", "w", encoding="utf-8") as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Erro ao salvar configurações: {e}")
    
    def center_window(self):
        """Centraliza a janela na tela"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")
    
    def create_widgets(self):
        """Cria a interface"""
        # Frame principal
        main_frame = ctk.CTkFrame(self.root)
        main_frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Título
        title_label = ctk.CTkLabel(
            main_frame, 
            text="CND Manager",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title_label.pack(pady=(15, 20))
        
        # Frame de seleção de pasta
        folder_frame = ctk.CTkFrame(main_frame)
        folder_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        folder_label = ctk.CTkLabel(folder_frame, text="Pasta Principal:", font=ctk.CTkFont(size=12, weight="bold"))
        folder_label.pack(anchor="w", padx=15, pady=(15, 5))
        
        folder_input_frame = ctk.CTkFrame(folder_frame)
        folder_input_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        self.folder_entry = ctk.CTkEntry(
            folder_input_frame, 
            textvariable=self.folder_path,
            font=ctk.CTkFont(size=11),
            height=32
        )
        self.folder_entry.pack(side="left", fill="x", expand=True, padx=(8, 8), pady=8)
        
        browse_btn = ctk.CTkButton(
            folder_input_frame,
            text="Procurar",
            command=self.browse_folder,
            width=80,
            height=32
        )
        browse_btn.pack(side="right", padx=(0, 8), pady=8)
        
        # Frame de configurações (compacto)
        config_frame = ctk.CTkFrame(main_frame)
        config_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        # Layout horizontal para configurações
        config_top_frame = ctk.CTkFrame(config_frame)
        config_top_frame.pack(fill="x", padx=15, pady=15)
        
        # Lado esquerdo - Tipos de arquivo
        left_config = ctk.CTkFrame(config_top_frame)
        left_config.pack(side="left", fill="both", expand=True, padx=(0, 8))
        
        files_label = ctk.CTkLabel(left_config, text="Tipos de CND:", font=ctk.CTkFont(size=11, weight="bold"))
        files_label.pack(anchor="w", padx=10, pady=(10, 5))
        
        self.files_text = ctk.CTkTextbox(left_config, height=60, font=ctk.CTkFont(size=10))
        self.files_text.pack(fill="x", padx=10, pady=(0, 10))
        self.files_text.insert("1.0", "\n".join(self.config["expected_files"]))
        
        # Lado direito - Texto alvo
        right_config = ctk.CTkFrame(config_top_frame)
        right_config.pack(side="right", fill="both", expand=True, padx=(8, 0))
        
        target_label = ctk.CTkLabel(right_config, text="Texto para Certidão Positiva:", font=ctk.CTkFont(size=11, weight="bold"))
        target_label.pack(anchor="w", padx=10, pady=(10, 5))
        
        self.target_entry = ctk.CTkEntry(right_config, font=ctk.CTkFont(size=10), height=28)
        self.target_entry.pack(fill="x", padx=10, pady=(0, 15))
        self.target_entry.insert(0, self.config["target_line"])
        
        # Frame de botões e progresso (compacto)
        control_frame = ctk.CTkFrame(main_frame)
        control_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        # Botões em linha
        buttons_frame = ctk.CTkFrame(control_frame)
        buttons_frame.pack(fill="x", padx=15, pady=(15, 10))
        
        self.process_btn = ctk.CTkButton(
            buttons_frame,
            text="Processar CNDs",
            command=self.start_processing,
            width=120,
            height=32,
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.process_btn.pack(side="left", padx=(0, 10))
        
        self.export_btn = ctk.CTkButton(
            buttons_frame,
            text="Exportar Excel",
            command=self.export_report,
            width=120,
            height=32,
            state="disabled"
        )
        self.export_btn.pack(side="left", padx=10)
        
        self.clear_btn = ctk.CTkButton(
            buttons_frame,
            text="Limpar Logs",
            command=self.clear_logs,
            width=100,
            height=32,
            fg_color="transparent",
            border_width=2
        )
        self.clear_btn.pack(side="left", padx=(10, 0))
        
        # Progresso
        progress_inner_frame = ctk.CTkFrame(control_frame)
        progress_inner_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        self.progress_label = ctk.CTkLabel(progress_inner_frame, text="Pronto para processar", font=ctk.CTkFont(size=11))
        self.progress_label.pack(pady=(10, 5))
        
        self.progress_bar = ctk.CTkProgressBar(progress_inner_frame, height=16)
        self.progress_bar.pack(fill="x", padx=15, pady=(0, 10))
        self.progress_bar.set(0)
        
        # Frame de resultados
        results_frame = ctk.CTkFrame(main_frame)
        results_frame.pack(fill="both", expand=True, padx=15)
        
        results_label = ctk.CTkLabel(results_frame, text="Resultados:", font=ctk.CTkFont(size=12, weight="bold"))
        results_label.pack(anchor="w", padx=15, pady=(15, 8))
        
        # Treeview para resultados
        tree_frame = ctk.CTkFrame(results_frame)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        # Criar Treeview
        columns = ("Empresa", "Municipal", "RFB", "FGTS", "PROC", "Estadual", "Positiva", "Status")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=10)
        
        # Configurar colunas
        self.tree.heading("Empresa", text="Empresa")
        self.tree.heading("Municipal", text="Municipal")
        self.tree.heading("RFB", text="RFB")
        self.tree.heading("FGTS", text="FGTS")
        self.tree.heading("PROC", text="PROC")
        self.tree.heading("Estadual", text="Estadual")
        self.tree.heading("Positiva", text="Cert. Positiva")
        self.tree.heading("Status", text="Status")
        
        # Configurar largura das colunas (ajustado para tela menor)
        self.tree.column("Empresa", width=180)
        self.tree.column("Municipal", width=70)
        self.tree.column("RFB", width=70)
        self.tree.column("FGTS", width=70)
        self.tree.column("PROC", width=70)
        self.tree.column("Estadual", width=70)
        self.tree.column("Positiva", width=90)
        self.tree.column("Status", width=90)
        
        # Scrollbar para o Treeview
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=8)
        scrollbar.pack(side="right", fill="y", pady=8)
        
        # Configurar estilo do Treeview
        style = ttk.Style()
        style.configure("Treeview", rowheight=22)
        style.configure("Treeview.Heading", font=('Helvetica', 9, 'bold'))
        
        # Definir a pasta inicial se existir
        if self.config["last_folder"] and os.path.exists(self.config["last_folder"]):
            self.folder_path.set(self.config["last_folder"])
    
    def browse_folder(self):
        """Abre diálogo para selecionar pasta"""
        folder = filedialog.askdirectory(title="Selecione a pasta principal com as CNDs")
        if folder:
            self.folder_path.set(folder)
            self.config["last_folder"] = folder
            self.save_config()
    
    def update_config(self):
        """Atualiza configurações baseadas na interface"""
        # Atualizar tipos de arquivo
        files_text = self.files_text.get("1.0", "end-1c")
        self.config["expected_files"] = [line.strip() for line in files_text.split("\n") if line.strip()]
        
        # Atualizar linha alvo
        self.config["target_line"] = self.target_entry.get().strip()
        
        self.save_config()
    
    def start_processing(self):
        """Inicia o processamento em thread separada"""
        if self.processing:
            return
        
        folder = self.folder_path.get().strip()
        if not folder:
            messagebox.showerror("Erro", "Por favor, selecione uma pasta!")
            return
        
        if not os.path.exists(folder):
            messagebox.showerror("Erro", "Pasta não encontrada!")
            return
        
        self.update_config()
        self.processing = True
        self.process_btn.configure(text="Processando...", state="disabled")
        self.export_btn.configure(state="disabled")
        
        # Limpar resultados anteriores
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Iniciar processamento em thread separada
        thread = threading.Thread(target=self.process_folder, args=(folder,))
        thread.daemon = True
        thread.start()
    
    def process_folder(self, main_folder):
        """Processa a pasta principal"""
        try:
            expected_files = self.config["expected_files"]
            target_line = self.config["target_line"]
            
            # Obter lista de subpastas
            subfolders = [f for f in os.listdir(main_folder) 
                         if os.path.isdir(os.path.join(main_folder, f))]
            
            total_folders = len(subfolders)
            if total_folders == 0:
                self.root.after(0, lambda: self.update_progress("Nenhuma subpasta encontrada!", 0))
                self.processing_complete()
                return
            
            self.results_data = []
            
            for i, subfolder in enumerate(subfolders):
                subfolder_path = os.path.join(main_folder, subfolder)
                
                # Atualizar progresso
                progress = (i + 1) / total_folders
                self.root.after(0, lambda p=progress, f=subfolder: self.update_progress(f"Processando: {f}", p))
                
                # Processar subpasta
                result = self.process_subfolder(subfolder_path, subfolder, expected_files, target_line)
                self.results_data.append(result)
                
                # Atualizar interface
                self.root.after(0, lambda r=result: self.add_result_to_tree(r))
            
            # Processamento concluído
            self.root.after(0, lambda: self.update_progress("Processamento concluído!", 1.0))
            self.processing_complete()
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Erro", f"Erro durante o processamento: {str(e)}"))
            self.processing_complete()
    
    def process_subfolder(self, subfolder_path, subfolder_name, expected_files, target_line):
        """Processa uma subpasta individual"""
        found_files = {file_type: False for file_type in expected_files}
        has_positive_cert = False
        
        try:
            # Procurar arquivos PDF
            for file_name in os.listdir(subfolder_path):
                if file_name.lower().endswith('.pdf'):
                    for file_type in expected_files:
                        if file_type in file_name.upper():
                            found_files[file_type] = True
                            
                            # Verificar certidão positiva especificamente para CND ESTADUAL
                            if file_type == "CND ESTADUAL":
                                pdf_path = os.path.join(subfolder_path, file_name)
                                if self.check_positive_cert(pdf_path, target_line):
                                    has_positive_cert = True
                                    self.create_log(pdf_path, target_line)
            
            # Verificar arquivos faltantes
            missing_files = [file_type for file_type, found in found_files.items() if not found]
            if missing_files:
                self.create_missing_files_log(subfolder_path, missing_files)
            
            # Preparar resultado
            return {
                "empresa": subfolder_name,
                "municipal": "SIM" if found_files.get("CND MUNICIPAL", False) else "NÃO",
                "rfb": "SIM" if found_files.get("CND RFB", False) else "NÃO",
                "fgts": "SIM" if found_files.get("CND FGTS", False) else "NÃO",
                "proc": "SIM" if found_files.get("CND PROC", False) else "NÃO",
                "estadual": "SIM" if found_files.get("CND ESTADUAL", False) else "NÃO",
                "positiva": "SIM" if has_positive_cert else "NÃO",
                "status": "COMPLETO" if not missing_files else "INCOMPLETO",
                "missing_files": missing_files
            }
            
        except Exception as e:
            return {
                "empresa": subfolder_name,
                "municipal": "ERRO",
                "rfb": "ERRO",
                "fgts": "ERRO",
                "proc": "ERRO",
                "estadual": "ERRO",
                "positiva": "ERRO",
                "status": "ERRO",
                "missing_files": []
            }
    
    def check_positive_cert(self, file_path, target_line):
        """Verifica se o PDF contém certidão positiva"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                
                for page in pdf_reader.pages:
                    page_text = page.extract_text()
                    if page_text and target_line in page_text:
                        return True
                return False
        except Exception as e:
            print(f"Erro ao ler PDF {file_path}: {str(e)}")
            return False
    
    def create_log(self, file_path, target_line):
        """Cria log para certidões positivas"""
        try:
            log_file = "pdf_log.txt"
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_message = f"[{timestamp}] Arquivo: {file_path}, Linha encontrada: {target_line}\n"
            
            with open(log_file, 'a', encoding='utf-8') as log:
                log.write(log_message)
        except Exception as e:
            print(f"Erro ao criar log: {e}")
    
    def create_missing_files_log(self, subfolder_path, missing_files):
        """Cria log para arquivos faltantes"""
        try:
            log_file = "missing_files_log.txt"
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_message = f"[{timestamp}] Subpasta: {subfolder_path}, Arquivos faltando: {', '.join(missing_files)}\n"
            
            with open(log_file, 'a', encoding='utf-8') as log:
                log.write(log_message)
        except Exception as e:
            print(f"Erro ao criar log de arquivos faltantes: {e}")
    
    def update_progress(self, text, value):
        """Atualiza a barra de progresso"""
        self.progress_label.configure(text=text)
        self.progress_bar.set(value)
    
    def add_result_to_tree(self, result):
        """Adiciona resultado ao Treeview"""
        values = (
            result["empresa"],
            result["municipal"],
            result["rfb"],
            result["fgts"],
            result["proc"],
            result["estadual"],
            result["positiva"],
            result["status"]
        )
        
        item = self.tree.insert("", "end", values=values)
        
        # Colorir linha baseado no status
        if result["status"] == "COMPLETO":
            self.tree.set(item, "Status", "✓ COMPLETO")
        elif result["status"] == "INCOMPLETO":
            self.tree.set(item, "Status", "⚠ INCOMPLETO")
        elif result["status"] == "ERRO":
            self.tree.set(item, "Status", "✗ ERRO")
        
        # Destacar certidões positivas
        if result["positiva"] == "SIM":
            self.tree.set(item, "Positiva", "⚠ SIM")
    
    def processing_complete(self):
        """Finaliza o processamento"""
        self.processing = False
        self.process_btn.configure(text="Processar CNDs", state="normal")
        if self.results_data:
            self.export_btn.configure(state="normal")
    
    def export_report(self):
        """Exporta relatório para Excel"""
        if not self.results_data:
            messagebox.showwarning("Aviso", "Nenhum dado para exportar!")
            return
        
        try:
            # Solicitar local para salvar
            filename = filedialog.asksaveasfilename(
                title="Salvar Relatório",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not filename:
                return
            
            self.create_excel_report(self.results_data, filename)
            messagebox.showinfo("Sucesso", f"Relatório exportado com sucesso!\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar relatório: {str(e)}")
    
    def create_excel_report(self, data, filename):
        """Cria relatório Excel com formatação"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório CND"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Cabeçalhos
        headers = ["Empresa", "CND MUNICIPAL", "CND RFB", "CND FGTS", "CND PROC", 
                  "CND ESTADUAL", "Certidão Positiva", "Arquivos Faltantes", "Status"]
        
        # Inserir título
        ws.merge_cells("A1:I1")
        title_cell = ws.cell(row=1, column=1, value=f"RELATÓRIO DE CONTROLE DE CND - {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = center_alignment
        
        # Adicionar cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Adicionar dados
        for row_idx, result in enumerate(data, 3):
            row_data = [
                result["empresa"],
                result["municipal"],
                result["rfb"],
                result["fgts"],
                result["proc"],
                result["estadual"],
                result["positiva"],
                ", ".join(result["missing_files"]) if result["missing_files"] else "NENHUM",
                result["status"]
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = center_alignment
                
                # Colorir células
                if col_idx == 9:  # Status
                    if value == "COMPLETO":
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif value == "INCOMPLETO":
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                elif col_idx == 7:  # Certidão Positiva
                    if value == "SIM":
                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
        
        # Ajustar larguras
        column_widths = [25, 15, 15, 15, 15, 15, 18, 30, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = width
        
        wb.save(filename)
    
    def clear_logs(self):
        """Limpa arquivos de log"""
        try:
            log_files = ["pdf_log.txt", "missing_files_log.txt"]
            deleted_files = []
            
            for log_file in log_files:
                if os.path.exists(log_file):
                    os.remove(log_file)
                    deleted_files.append(log_file)
            
            if deleted_files:
                messagebox.showinfo("Sucesso", f"Logs limpos: {', '.join(deleted_files)}")
            else:
                messagebox.showinfo("Info", "Nenhum arquivo de log encontrado.")
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao limpar logs: {str(e)}")
    
    def run(self):
        """Executa a aplicação"""
        self.root.mainloop()

if __name__ == "__main__":
    app = CNDManager()
    app.run()