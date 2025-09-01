import re
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import PyPDF2
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import threading
import json
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

class CNDManager:
    def __init__(self):
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        
        self.root = ctk.CTk()
        self.root.title("CND Manager - Controle de Certidões")
        self.root.geometry("1000x700")
        self.root.minsize(900, 600)
        
        self.folder_path = tk.StringVar()
        self.processing = False
        self.results_data = []
        
        self.config = {
            "expected_files": ["CND MUNICIPAL", "CND RFB", "CND FGTS", "CND PROC", "CND ESTADUAL"],
            "target_line": "CERTIDÃO POSITIVA DE DÉBITOS - CPD",
            "last_folder": "",
            "mode": "Verificar Positiva"
        }
        
        self.load_config()
        self.create_widgets()
        self.center_window()

    def load_config(self):
        try:
            if os.path.exists("cnd_config.json"):
                with open("cnd_config.json", "r", encoding="utf-8") as f:
                    loaded_config = json.load(f)
                    self.config.update(loaded_config)
        except Exception as e:
            print(f"Erro ao carregar configurações: {e}")

    def save_config(self):
        try:
            with open("cnd_config.json", "w", encoding="utf-8") as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Erro ao salvar configurações: {e}")

    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def create_widgets(self):
        main_frame = ctk.CTkFrame(self.root)
        main_frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        title_label = ctk.CTkLabel(main_frame, text="CND Manager",
                                   font=ctk.CTkFont(size=20, weight="bold"))
        title_label.pack(pady=(15, 20))
        
        folder_frame = ctk.CTkFrame(main_frame)
        folder_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        folder_label = ctk.CTkLabel(folder_frame, text="Pasta Principal:", font=ctk.CTkFont(size=12, weight="bold"))
        folder_label.pack(anchor="w", padx=15, pady=(15, 5))
        
        folder_input_frame = ctk.CTkFrame(folder_frame)
        folder_input_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        self.folder_entry = ctk.CTkEntry(folder_input_frame, textvariable=self.folder_path,
                                         font=ctk.CTkFont(size=11), height=32)
        self.folder_entry.pack(side="left", fill="x", expand=True, padx=(8, 8), pady=8)
        
        browse_btn = ctk.CTkButton(folder_input_frame, text="Procurar", command=self.browse_folder,
                                   width=80, height=32)
        browse_btn.pack(side="right", padx=(0, 8), pady=8)
        
        # Escolha do modo
        mode_frame = ctk.CTkFrame(main_frame)
        mode_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        mode_label = ctk.CTkLabel(mode_frame, text="Modo de Verificação:", font=ctk.CTkFont(size=12, weight="bold"))
        mode_label.pack(side="left", padx=10, pady=10)
        
        self.mode_var = tk.StringVar(value=self.config["mode"])
        self.mode_combo = ttk.Combobox(mode_frame, textvariable=self.mode_var, state="readonly",
                                       values=["Verificar Positiva", "Verificar Vencimento"])
        self.mode_combo.pack(side="left", padx=10, pady=10)
        
        control_frame = ctk.CTkFrame(main_frame)
        control_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        buttons_frame = ctk.CTkFrame(control_frame)
        buttons_frame.pack(fill="x", padx=15, pady=(15, 10))
        
        self.process_btn = ctk.CTkButton(buttons_frame, text="Processar", command=self.start_processing,
                                         width=120, height=32, font=ctk.CTkFont(size=12, weight="bold"))
        self.process_btn.pack(side="left", padx=(0, 10))
        
        self.export_btn = ctk.CTkButton(buttons_frame, text="Exportar Excel",
                                        command=self.export_report, width=120, height=32,
                                        state="disabled")
        self.export_btn.pack(side="left", padx=10)
        
        self.progress_label = ctk.CTkLabel(control_frame, text="Pronto para processar", font=ctk.CTkFont(size=11))
        self.progress_label.pack(pady=(10, 5))
        
        self.progress_bar = ctk.CTkProgressBar(control_frame, height=16)
        self.progress_bar.pack(fill="x", padx=15, pady=(0, 10))
        self.progress_bar.set(0)
        
        results_frame = ctk.CTkFrame(main_frame)
        results_frame.pack(fill="both", expand=True, padx=15)
        
        results_label = ctk.CTkLabel(results_frame, text="Resultados:", font=ctk.CTkFont(size=12, weight="bold"))
        results_label.pack(anchor="w", padx=15, pady=(15, 8))
        
        tree_frame = ctk.CTkFrame(results_frame)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        self.tree = ttk.Treeview(tree_frame, show="headings", height=10)
        self.tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=8)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y", pady=8)

    def browse_folder(self):
        folder = filedialog.askdirectory(title="Selecione a pasta principal com as CNDs")
        if folder:
            self.folder_path.set(folder)
            self.config["last_folder"] = folder
            self.save_config()

    def start_processing(self):
        if self.processing:
            return
        folder = self.folder_path.get().strip()
        if not folder or not os.path.exists(folder):
            messagebox.showerror("Erro", "Por favor, selecione uma pasta válida!")
            return

        self.config["mode"] = self.mode_var.get()
        self.save_config()

        self.processing = True
        self.process_btn.configure(text="Processando...", state="disabled")
        self.export_btn.configure(state="disabled")

        for item in self.tree.get_children():
            self.tree.delete(item)

        thread = threading.Thread(target=self.process_folder, args=(folder,))
        thread.daemon = True
        thread.start()

    def process_folder(self, main_folder):
        try:
            mode = self.config["mode"]
            expected_files = self.config["expected_files"]
            target_line = self.config["target_line"]

            subfolders = [f for f in os.listdir(main_folder) if os.path.isdir(os.path.join(main_folder, f))]
            total_folders = len(subfolders)
            if total_folders == 0:
                self.root.after(0, lambda: self.update_progress("Nenhuma subpasta encontrada!", 0))
                self.processing_complete()
                return

            self.results_data = []

            for i, subfolder in enumerate(subfolders):
                subfolder_path = os.path.join(main_folder, subfolder)
                progress = (i + 1) / total_folders
                self.root.after(0, lambda p=progress, f=subfolder: self.update_progress(f"Processando: {f}", p))

                if mode == "Verificar Positiva":
                    result = self.process_subfolder_positive(subfolder_path, subfolder, expected_files, target_line)
                else:
                    result = self.process_subfolder_vencimento(subfolder_path, subfolder, expected_files)

                self.results_data.append(result)
                self.root.after(0, lambda r=result: self.add_result_to_tree(r))

            self.root.after(0, lambda: self.update_progress("Processamento concluído!", 1.0))
            self.processing_complete()
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Erro", str(e)))
            self.processing_complete()

    # --- Modo POSITIVA ---
    def process_subfolder_positive(self, subfolder_path, subfolder_name, expected_files, target_line):
        found_files = {file_type: False for file_type in expected_files}
        positive_cert_type = None  # <- novo
        try:
            for file_name in os.listdir(subfolder_path):
                if file_name.lower().endswith('.pdf'):
                    for file_type in expected_files:
                        if file_type in file_name.upper():
                            found_files[file_type] = True
                            pdf_path = os.path.join(subfolder_path, file_name)
                            if self.check_positive_cert(pdf_path, target_line):
                                positive_cert_type = file_type  # <- guarda qual certidão
            missing_files = [f for f, found in found_files.items() if not found]
            return {
                "empresa": subfolder_name,
                "municipal": "SIM" if found_files["CND MUNICIPAL"] else "NÃO",
                "rfb": "SIM" if found_files["CND RFB"] else "NÃO",
                "fgts": "SIM" if found_files["CND FGTS"] else "NÃO",
                "proc": "SIM" if found_files["CND PROC"] else "NÃO",
                "estadual": "SIM" if found_files["CND ESTADUAL"] else "NÃO",
                "positiva": positive_cert_type if positive_cert_type else "NENHUMA",  # <- mudança
                "status": "COMPLETO" if not missing_files else "INCOMPLETO",
                "missing_files": missing_files
            }
        except:
            return {"empresa": subfolder_name, "status": "ERRO"}

    def check_positive_cert(self, file_path, target_line):
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    page_text = page.extract_text()
                    if page_text and target_line in page_text:
                        return True
            return False
        except:
            return False

    # --- Modo VENCIMENTO ---
    def check_due_date(self, file_name):
        try:
            # procura a data no formato dd.mm.yyyy
            match = re.search(r"\d{2}\.\d{2}\.\d{4}", file_name)
            if not match:
                return "DATA NÃO ENCONTRADA"
            
            date_str = match.group()
            due_date = datetime.strptime(date_str, "%d.%m.%Y").date()
            today = datetime.today().date()
            
            return "VENCIDA" if due_date < today else "VÁLIDA"
        except Exception as e:
            return "ERRO DATA"

    def process_subfolder_vencimento(self, subfolder_path, subfolder_name, expected_files):
        found_files = {file_type: "NÃO" for file_type in expected_files}
        try:
            for file_name in os.listdir(subfolder_path):
                if file_name.lower().endswith('.pdf'):
                    for file_type in expected_files:
                        if file_type in file_name.upper():
                            found_files[file_type] = self.check_due_date(file_name)
            missing_files = [f for f, status in found_files.items() if status == "NÃO"]
            return {
                "empresa": subfolder_name,
                "municipal": found_files["CND MUNICIPAL"],
                "rfb": found_files["CND RFB"],
                "fgts": found_files["CND FGTS"],
                "proc": found_files["CND PROC"],
                "estadual": found_files["CND ESTADUAL"],
                "status": "COMPLETO" if not missing_files else "INCOMPLETO",
                "missing_files": missing_files
            }
        except:
            return {"empresa": subfolder_name, "status": "ERRO"}

    def update_progress(self, text, value):
        self.progress_label.configure(text=text)
        self.progress_bar.set(value)

    def add_result_to_tree(self, result):
        # Definir colunas dependendo do modo
        mode = self.config["mode"]
        if mode == "Verificar Positiva":
            columns = ("Empresa", "Municipal", "RFB", "FGTS", "PROC", "Estadual", "Positiva", "Status")
            self.tree["columns"] = columns
            for col in columns:
                self.tree.heading(col, text=col)
                self.tree.column(col, width=100)
            values = (result.get("empresa",""), result.get("municipal",""), result.get("rfb",""),
                      result.get("fgts",""), result.get("proc",""), result.get("estadual",""),
                      result.get("positiva",""), result.get("status",""))
        else:
            columns = ("Empresa", "Municipal", "RFB", "FGTS", "PROC", "Estadual", "Status")
            self.tree["columns"] = columns
            for col in columns:
                self.tree.heading(col, text=col)
                self.tree.column(col, width=100)
            values = (result.get("empresa",""), result.get("municipal",""), result.get("rfb",""),
                      result.get("fgts",""), result.get("proc",""), result.get("estadual",""),
                      result.get("status",""))
        self.tree.insert("", "end", values=values)

    def processing_complete(self):
        self.processing = False
        self.process_btn.configure(text="Processar", state="normal")
        if self.results_data:
            self.export_btn.configure(state="normal")

    def export_report(self):
        if not self.results_data:
            messagebox.showwarning("Aviso", "Nenhum dado para exportar!")
            return
        try:
            filename = filedialog.asksaveasfilename(
                title="Salvar Relatório",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not filename:
                return
            self.create_excel_report(self.results_data, filename)
            messagebox.showinfo("Sucesso", f"Relatório exportado com sucesso!\n{filename}")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def create_excel_report(self, data, filename):
        wb = Workbook()
        ws = wb.active
        mode = self.config["mode"]
        ws.title = "Relatório CND"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))

        if mode == "Verificar Positiva":
            headers = ["Empresa", "CND MUNICIPAL", "CND RFB", "CND FGTS", "CND PROC", 
                    "CND ESTADUAL", "Certidão Positiva", "Arquivos Faltantes", "Status"]
        else:
            headers = ["Empresa", "CND MUNICIPAL", "CND RFB", "CND FGTS", "CND PROC", 
                    "CND ESTADUAL", "Arquivos Faltantes", "Status"]
        
        # título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1,
                            value=f"RELATÓRIO DE CND ({mode}) - {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = center_alignment
        
        # cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # dados
        for row_idx, result in enumerate(data, 3):
            if mode == "Verificar Positiva":
                row_data = [result["empresa"], result["municipal"], result["rfb"], result["fgts"],
                            result["proc"], result["estadual"],
                            ", ".join(result["missing_files"]) if result["missing_files"] else "NENHUM",
                            result["status"]]
            else:
                row_data = [result["empresa"], result["municipal"], result["rfb"], result["fgts"],
                            result["proc"], result["estadual"],
                            ", ".join(result["missing_files"]) if result["missing_files"] else "NENHUM",
                            result["status"]]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = center_alignment
        
        # opções do dropdown
        opcoes = ["VENCIDA", "VÁLIDA", "IMPEDIDA", "TAREFA", "NÃO"]

        # colunas que devem receber validação (todas de status de CND + Status final)
        if mode == "Verificar Positiva":
            colunas_status = [2, 3, 4, 5, 6, 9]  # municipal, rfb, fgts, proc, estadual, status final
        else:
            colunas_status = [2, 3, 4, 5, 6, 8]  # municipal, rfb, fgts, proc, estadual, status final
        
        # aplicar dropdown e cores em cada coluna de status
        cores = {
            "VENCIDA": "FF0000",   # vermelho
            "VÁLIDA": "00FF00",    # verde
            "IMPEDIDA": "FFA500",  # laranja
            "TAREFA": "0000FF",    # azul
            "NÃO": "808080"        # cinza
        }
        
        for col in colunas_status:
            col_letter = chr(64 + col)  # converte número da coluna em letra (2 -> B, etc.)
            
            # validação dropdown
            dv = DataValidation(type="list", formula1=f'"{",".join(opcoes)}"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}3:{col_letter}500")
            
            # formatação condicional
            for status, cor in cores.items():
                formula = f'EXACT("{status}",${col_letter}3)'
                rule = FormulaRule(formula=[formula],
                                fill=PatternFill(start_color=cor, end_color=cor, fill_type="solid"))
                ws.conditional_formatting.add(f"{col_letter}3:{col_letter}500", rule)

        wb.save(filename)

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = CNDManager()
    app.run()
